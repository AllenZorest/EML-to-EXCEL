#!/usr/bin/env python3
"""
EML to Excel — Streamlit Web 界面
=====================================
上传 .eml 文件 → 实时解析预览 → 一键下载 Excel

依赖：streamlit, openpyxl, chardet, pandas
"""

import io
import re
import tempfile
from datetime import datetime, timezone, timedelta
from email import policy
from email.parser import BytesParser
from email.header import decode_header, make_header
from email.utils import parsedate_to_datetime, getaddresses
from pathlib import Path
from typing import Optional

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─── 常量 ─────────────────────────────────────────────
CST = timezone(timedelta(hours=8))
HEADERS = ["file", "from", "date", "to", "cc", "subject", "body"]
COL_WIDTHS = [55, 38, 22, 50, 38, 60, 100]

COL_LABELS = {
    "file": "文件名",
    "from": "发件人",
    "date": "时间",
    "to": "收件人",
    "cc": "抄送",
    "subject": "标题",
    "body": "正文",
}

# ─── 解析逻辑（复用 eml_to_excel.py 核心）────────────────

def decode_mime_header(raw: str) -> str:
    if not raw:
        return ""
    try:
        return str(make_header(decode_header(raw)))
    except Exception:
        return raw


def extract_addresses(raw: str) -> str:
    if not raw:
        return ""
    pairs = getaddresses([raw])
    return ", ".join(addr.strip().lower() for _, addr in pairs if addr)


def _decode_part(part) -> str:
    try:
        charset = part.get_content_charset() or "utf-8"
        payload = part.get_payload(decode=True)
        if payload is None:
            return ""
        return payload.decode(charset, errors="replace")
    except (LookupError, UnicodeDecodeError):
        payload = part.get_payload(decode=True)
        if payload is None:
            return ""
        try:
            import chardet
            result = chardet.detect(payload)
            encoding = result.get("encoding") or "utf-8"
            return payload.decode(encoding, errors="replace")
        except ImportError:
            return payload.decode("utf-8", errors="replace")


def _strip_html(html: str) -> str:
    html = re.sub(r'<(style|script)[^>]*>.*?</\1>', '', html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', html)
    for ent, rep in [("&nbsp;", " "), ("&amp;", "&"), ("&lt;", "<"),
                     ("&gt;", ">"), ("&quot;", '"'), ("&#39;", "'")]:
        text = text.replace(ent, rep)
    return re.sub(r'\s+', ' ', text).strip()


def extract_body(msg) -> str:
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                return _decode_part(part)
        for part in msg.walk():
            if part.get_content_type() == "text/html":
                return _strip_html(_decode_part(part))
        return ""
    return _decode_part(msg)


def _parse_date_fallback(date_str: str) -> str:
    if not date_str:
        return ""
    for fmt in ["%a, %d %b %Y %H:%M:%S %z", "%a, %d %b %Y %H:%M:%S",
                "%d %b %Y %H:%M:%S %z", "%Y-%m-%d %H:%M:%S"]:
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(CST).strftime("%Y/%m/%d %H:%M:%S")
        except ValueError:
            continue
    return date_str.strip()


def parse_eml_bytes(filename: str, data: bytes) -> Optional[dict]:
    """从字节流解析一封 EML"""
    try:
        msg = BytesParser(policy=policy.default).parsebytes(data)
    except Exception as e:
        return None

    date_raw = msg.get("Date", "")
    try:
        dt = parsedate_to_datetime(date_raw)
        date_str = dt.astimezone(CST).strftime("%Y/%m/%d %H:%M:%S")
    except Exception:
        date_str = _parse_date_fallback(date_raw)

    body = extract_body(msg)[:8000]

    return {
        "file": filename,
        "from": extract_addresses(msg.get("From", "")),
        "date": date_str,
        "to": extract_addresses(msg.get("To", "")),
        "cc": extract_addresses(msg.get("Cc", "")),
        "subject": decode_mime_header(msg.get("Subject", "")),
        "body": body,
    }


# ─── Excel 导出 ───────────────────────────────────────

def build_excel(rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"

    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    data_font = Font(name="Microsoft YaHei", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    date_align = Alignment(horizontal="center", vertical="top")

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=COL_LABELS[h])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    for col_idx, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, record in enumerate(rows, 2):
        for col_idx, key in enumerate(HEADERS, 1):
            value = record.get(key) or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin
            cell.alignment = date_align if key == "date" else data_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Streamlit UI ─────────────────────────────────────

def main():
    st.set_page_config(
        page_title="EML → Excel | Nexas Email Parser",
        page_icon="📧",
        layout="wide",
    )

    # ── 顶部标题区 ──
    st.markdown("""
    <div style='background: linear-gradient(90deg, #4472C4, #2E5BA8); padding: 20px 28px; border-radius: 12px; margin-bottom: 24px;'>
        <h1 style='color: white; margin: 0; font-size: 26px;'>📧 EML → Excel 邮件解析工具</h1>
        <p style='color: #cdd9f5; margin: 6px 0 0; font-size: 14px;'>为日本 Nexas 字体公司开发 · 支持中/日/英多语种 · 自动处理 MIME 编码与时区</p>
    </div>
    """, unsafe_allow_html=True)

    # ── 侧边栏 ──
    with st.sidebar:
        st.markdown("### ⚙️ 设置")
        body_preview_len = st.slider("正文预览字数", 50, 500, 200, step=50)
        show_body = st.checkbox("在表格中显示正文", value=False)
        st.markdown("---")
        st.markdown("""
        **使用说明**
        1. 点击「上传 EML 文件」
        2. 支持同时上传多个文件
        3. 解析完成后点击「下载 Excel」

        **支持格式**
        - 标准 RFC 2822 EML
        - Base64 / Quoted-Printable 编码
        - multipart 附件邮件
        - 中/日/英 多语种标头

        **时区**：统一转为北京时间 UTC+8
        """)
        st.markdown("---")
        st.markdown(
            "[![GitHub](https://img.shields.io/badge/GitHub-AllenZorest-black?logo=github)](https://github.com/AllenZorest/EML-to-EXCEL)",
            unsafe_allow_html=True,
        )

    # ── 上传区 ──
    uploaded = st.file_uploader(
        "上传 EML 文件（支持批量）",
        type=["eml"],
        accept_multiple_files=True,
        help="从邮件客户端导出的 .eml 格式文件，可一次上传多个"
    )

    if not uploaded:
        st.info("👆 请上传 .eml 文件开始解析")

        # 展示功能说明卡片
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("""
            <div style='border: 1px solid #e0e0e0; border-radius: 10px; padding: 16px; text-align: center;'>
                <div style='font-size: 32px;'>🔐</div>
                <strong>MIME 编码解码</strong>
                <p style='color: #666; font-size: 13px; margin-top: 8px;'>自动处理 Base64、Quoted-Printable、RFC 2047 标头编码</p>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown("""
            <div style='border: 1px solid #e0e0e0; border-radius: 10px; padding: 16px; text-align: center;'>
                <div style='font-size: 32px;'>🌐</div>
                <strong>多时区归一化</strong>
                <p style='color: #666; font-size: 13px; margin-top: 8px;'>将 +0300/+0900 等时区统一转为北京时间 UTC+8</p>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown("""
            <div style='border: 1px solid #e0e0e0; border-radius: 10px; padding: 16px; text-align: center;'>
                <div style='font-size: 32px;'>📊</div>
                <strong>结构化输出</strong>
                <p style='color: #666; font-size: 13px; margin-top: 8px;'>带冻结表头、自动筛选的格式化 Excel，可直接检索</p>
            </div>
            """, unsafe_allow_html=True)
        return

    # ── 解析进度 ──
    results = []
    failed_files = []

    progress_bar = st.progress(0, text="准备解析...")
    total = len(uploaded)

    for idx, f in enumerate(uploaded):
        data = f.read()
        record = parse_eml_bytes(f.name, data)
        if record:
            results.append(record)
        else:
            failed_files.append(f.name)

        pct = (idx + 1) / total
        progress_bar.progress(pct, text=f"正在解析 {idx + 1}/{total}：{f.name[:40]}...")

    progress_bar.empty()

    if not results:
        st.error("❌ 没有成功解析任何邮件，请检查文件格式。")
        return

    # ── 统计摘要 ──
    st.success(f"✅ 解析完成：{len(results)} 封成功，{len(failed_files)} 封失败")

    metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
    senders = len(set(r["from"] for r in results if r["from"]))
    recipients = len(set(
        addr.strip()
        for r in results
        for addr in r["to"].split(",") if addr.strip()
    ))

    dates = [r["date"] for r in results if r["date"]]
    date_range = f"{min(dates)[:10]} ~ {max(dates)[:10]}" if dates else "-"

    metric_col1.metric("📧 邮件总数", len(results))
    metric_col2.metric("👤 发件人数", senders)
    metric_col3.metric("📬 收件人数", recipients)
    metric_col4.metric("📅 时间跨度", date_range)

    # ── 数据预览 ──
    st.markdown("### 📋 解析结果预览")

    df = pd.DataFrame(results)
    # 正文截断显示
    if show_body:
        df_display = df.copy()
        df_display["body"] = df_display["body"].str[:body_preview_len] + "..."
    else:
        df_display = df.drop(columns=["body"])

    df_display.columns = [COL_LABELS.get(c, c) for c in df_display.columns]
    st.dataframe(df_display, use_container_width=True, height=400)

    # ── 失败文件列表 ──
    if failed_files:
        with st.expander(f"⚠️ {len(failed_files)} 个文件解析失败"):
            for name in failed_files:
                st.text(f"  • {name}")

    # ── 下载按钮 ──
    st.markdown("---")
    excel_bytes = build_excel(results)
    now_str = datetime.now(CST).strftime("%Y%m%d_%H%M%S")
    filename = f"emails_{now_str}.xlsx"

    col_dl, col_info = st.columns([1, 3])
    with col_dl:
        st.download_button(
            label="⬇️ 下载 Excel",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
    with col_info:
        st.caption(f"输出文件：{filename} · {len(results)} 行 · 含冻结表头与自动筛选")


if __name__ == "__main__":
    main()
