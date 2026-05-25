#!/usr/bin/env python3
"""
EML 邮件批量转 Excel 工具
============================
用途：递归扫描指定目录下所有 .eml 文件，提取邮件元信息与正文，输出为结构化 Excel 表格。

项目背景：为日本 Nexas 公司处理 1800+ 封商务邮件数据，实现邮件归档与检索自动化。

输出字段：
  file    — 文件名（不含路径）
  from    — 发信人邮箱地址
  date    — 发送时间，格式 YYYY/MM/DD HH:MM:SS
  to      — 收件人邮箱（多人用逗号分隔）
  cc      — 抄送邮箱（多人用逗号分隔）
  subject — 邮件标题
  body    — 邮件正文纯文本

依赖：Python 3.8+，openpyxl，chardet
用法：python eml_to_excel.py <EML目录> [输出Excel路径]
"""

import os
import sys
import re
import logging
from pathlib import Path
from datetime import datetime, timezone, timedelta
from email import policy
from email.parser import BytesParser
from email.header import decode_header, make_header
from email.utils import parsedate_to_datetime, getaddresses
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------- 日志配置 ----------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------- 常量 ----------
HEADERS = ["file", "from", "date", "to", "cc", "subject", "body"]
COL_WIDTHS = [55, 38, 22, 50, 38, 60, 100]  # 各列宽度
CST = timezone(timedelta(hours=8))  # 北京时间 (UTC+8)


# ===================== 邮件解析 =====================

def decode_mime_header(raw: str) -> str:
    """解码 MIME 编码的邮件标头（=?utf-8?B?...?= 等）"""
    if not raw:
        return ""
    try:
        fragments = decode_header(raw)
        return str(make_header(fragments))
    except Exception:
        # 兜底：尝试直接 decode
        return raw


def extract_addresses(raw: str) -> str:
    """从标头字符串中提取纯邮箱地址，逗号分隔"""
    if not raw:
        return ""
    pairs = getaddresses([raw])
    addrs = []
    for name, addr in pairs:
        if addr:
            addrs.append(addr.strip().lower())
    return ", ".join(addrs)


def extract_body(msg) -> str:
    """
    递归提取邮件正文。优先 text/plain，其次 text/html（去除标签）。
    处理 multipart/alternative, multipart/mixed 等类型。
    """
    if msg.is_multipart():
        # 先找 text/plain
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == "text/plain":
                return _decode_part(part)
        # 再找 text/html
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == "text/html":
                html = _decode_part(part)
                return _strip_html(html)
        # 都没找到
        return ""
    else:
        return _decode_part(msg)


def _decode_part(part) -> str:
    """解码单个 MIME part 的内容"""
    try:
        charset = part.get_content_charset() or "utf-8"
        payload = part.get_payload(decode=True)
        if payload is None:
            return ""
        return payload.decode(charset, errors="replace")
    except (LookupError, UnicodeDecodeError):
        # 编码声明不可靠，用 chardet 猜测
        payload = part.get_payload(decode=True)
        if payload is None:
            return ""
        try:
            import chardet
            result = chardet.detect(payload)
            encoding = result["encoding"] or "utf-8"
            return payload.decode(encoding, errors="replace")
        except ImportError:
            return payload.decode("utf-8", errors="replace")


def _strip_html(html: str) -> str:
    """简易 HTML 标签剥离，提取纯文本"""
    # 移除 style/script
    html = re.sub(r'<(style|script)[^>]*>.*?</\1>', '', html, flags=re.DOTALL | re.IGNORECASE)
    # 移除标签
    text = re.sub(r'<[^>]+>', '', html)
    # 处理常见实体
    text = text.replace("&nbsp;", " ").replace("&amp;", "&").replace("&lt;", "<")
    text = text.replace("&gt;", ">").replace("&quot;", '"').replace("&#39;", "'")
    # 合并空白
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def parse_eml(filepath: Path) -> Optional[dict]:
    """解析单个 EML 文件，返回字典或 None（解析失败）"""
    try:
        with open(filepath, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)
    except Exception as e:
        logger.warning(f"无法读取文件: {filepath.name} — {e}")
        return None

    # --- 发件人 ---
    from_raw = msg.get("From", "")
    from_addr = extract_addresses(from_raw)

    # --- 收件人 ---
    to_raw = msg.get("To", "")
    to_addr = extract_addresses(to_raw)

    # --- 抄送 ---
    cc_raw = msg.get("Cc", "")
    cc_addr = extract_addresses(cc_raw)

    # --- 标题 ---
    subject_raw = msg.get("Subject", "")
    subject = decode_mime_header(subject_raw)

    # --- 日期（统一转为北京时间 UTC+8）---
    date_raw = msg.get("Date", "")
    try:
        dt = parsedate_to_datetime(date_raw)
        dt_cst = dt.astimezone(CST)
        date_str = dt_cst.strftime("%Y/%m/%d %H:%M:%S")
    except Exception:
        date_str = _parse_date_fallback(date_raw)

    # --- 正文 ---
    body = extract_body(msg)
    # 限制正文长度（避免 Excel 单元格过载），保留前 8000 字符
    body = body[:8000]

    # --- 文件名 ---
    filename = str(filepath.name)  # 用原始字节文件名

    return {
        "file": filename,
        "from": from_addr,
        "date": date_str,
        "to": to_addr,
        "cc": cc_addr,
        "subject": subject,
        "body": body,
    }


def _parse_date_fallback(date_str: str) -> str:
    """日期解析兜底（统一转北京时间 UTC+8）"""
    if not date_str:
        return ""
    formats = [
        "%a, %d %b %Y %H:%M:%S %z",
        "%a, %d %b %Y %H:%M:%S",
        "%d %b %Y %H:%M:%S %z",
        "%a, %d %b %Y %H:%M:%S %Z",
        "%Y-%m-%d %H:%M:%S",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            dt_cst = dt.astimezone(CST)
            return dt_cst.strftime("%Y/%m/%d %H:%M:%S")
        except ValueError:
            continue
    return date_str.strip()


# ===================== Excel 输出 =====================

def write_excel(rows: list, output_path: str):
    """将解析结果写入 Excel，带格式化"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"

    # --- 表头样式 ---
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # --- 写入表头 ---
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- 设置列宽 ---
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- 数据样式 ---
    data_font = Font(name="Microsoft YaHei", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    date_align = Alignment(horizontal="center", vertical="top")

    # --- 写入数据 ---
    for row_idx, record in enumerate(rows, 2):
        for col_idx, key in enumerate(HEADERS, 1):
            value = record.get(key, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if key == "date":
                cell.alignment = date_align
            else:
                cell.alignment = data_align

    # --- 冻结首行 & 自动筛选 ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    # --- 保存 ---
    wb.save(output_path)
    logger.info(f"Excel 已保存: {output_path}")
    logger.info(f"共 {len(rows)} 条记录")


# ===================== 主流程 =====================

def main():
    if len(sys.argv) < 2:
        print("用法: python eml_to_excel.py <EML目录路径> [输出Excel路径]")
        print("示例: python eml_to_excel.py ./emails ./output.xlsx")
        sys.exit(1)

    eml_dir = Path(sys.argv[1])
    if not eml_dir.is_dir():
        logger.error(f"目录不存在: {eml_dir}")
        sys.exit(1)

    # 默认输出路径
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        output_path = eml_dir.parent / f"{eml_dir.name}_emails.xlsx"

    # --- 收集所有 EML 文件 ---
    eml_files = sorted(eml_dir.rglob("*.eml"))
    total = len(eml_files)
    logger.info(f"找到 {total} 个 EML 文件")
    logger.info(f"源目录: {eml_dir}")
    logger.info(f"输出路径: {output_path}")
    print()

    # --- 逐文件解析 ---
    results = []
    success = 0
    failed = 0

    for idx, fp in enumerate(eml_files, 1):
        record = parse_eml(fp)
        if record:
            results.append(record)
            success += 1
        else:
            failed += 1

        # 每 100 封打印进度
        if idx % 100 == 0 or idx == total:
            pct = idx / total * 100
            logger.info(
                f"进度: {idx}/{total} ({pct:.1f}%)  "
                f"成功={success}  失败={failed}"
            )

    # --- 输出 ---
    if not results:
        logger.error("没有成功解析任何邮件，请检查 EML 文件格式。")
        sys.exit(1)

    write_excel(results, str(output_path))

    # --- 汇总 ---
    print()
    logger.info("=" * 50)
    logger.info(f"处理完成!")
    logger.info(f"  总文件数: {total}")
    logger.info(f"  成功解析: {success}")
    logger.info(f"  解析失败: {failed}")
    logger.info(f"  输出文件: {output_path}")
    logger.info("=" * 50)


if __name__ == "__main__":
    main()
