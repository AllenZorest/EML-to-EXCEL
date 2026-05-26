#!/usr/bin/env python3
"""
Nexas EML 邮件解析器 — Streamlit Web 版
===========================================
基于 eml_to_excel.py 核心逻辑，提供 Web 交互界面：
  - 拖拽上传 .eml 文件
  - 实时解析预览
  - 数据表格（搜索/排序/筛选）
  - 统计仪表盘
  - 一键导出 Excel

启动：streamlit run streamlit_app.py
"""

import io
import re
import os
import logging
from datetime import datetime, timezone, timedelta
from email import policy
from email.parser import BytesParser
from email.header import decode_header, make_header
from email.utils import parsedate_to_datetime, getaddresses
from pathlib import Path
from typing import Optional
from collections import Counter

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import chardet

# ---------- 页面配置 ----------
st.set_page_config(
    page_title="Nexas EML 邮件解析器",
    page_icon="📧",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------- 常量 ----------
CST = timezone(timedelta(hours=8))
HEADERS = ["file", "from", "date", "to", "cc", "subject", "body"]


# ===================== 邮件解析逻辑（与 eml_to_excel.py 一致）=====================

def decode_mime_header(raw: str) -> str:
    if not raw:
        return ""
    try:
        fragments = decode_header(raw)
        return str(make_header(fragments))
    except Exception:
        return raw


def extract_addresses(raw: str) -> str:
    if not raw:
        return ""
    pairs = getaddresses([raw])
    addrs = [addr.strip().lower() for _, addr in pairs if addr]
    return ", ".join(addrs)


def extract_body(msg) -> str:
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                return _decode_part(part)
        for part in msg.walk():
            if part.get_content_type() == "text/html":
                html = _decode_part(part)
                return _strip_html(html)
        return ""
    else:
        return _decode_part(msg)


def _decode_part(part) -> str:
    try:
        charset = part.get_content_charset() or "utf-8"
        payload = part.get_payload(decode=True)
        if payload is None:
            return ""
        return payload.decode(charset, errors="replace")
    except (LookupError, UnicodeDecodeError):
        payload = part.get_payload(decode=True)
        if payload is None:
            return ""
        try:
            result = chardet.detect(payload)
            encoding = result["encoding"] or "utf-8"
            return payload.decode(encoding, errors="replace")
        except Exception:
            return payload.decode("utf-8", errors="replace")


def _strip_html(html: str) -> str:
    html = re.sub(r'<(style|script)[^>]*>.*?</\1>', '', html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', html)
    text = text.replace("&nbsp;", " ").replace("&amp;", "&").replace("&lt;", "<")
    text = text.replace("&gt;", ">").replace("&quot;", '"').replace("&#39;", "'")
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def parse_eml_bytes(filename: str, data: bytes) -> Optional[dict]:
    """解析 EML 字节数据"""
    try:
        msg = BytesParser(policy=policy.default).parse(io.BytesIO(data))
    except Exception:
        return None

    from_addr = extract_addresses(msg.get("From", ""))
    to_addr = extract_addresses(msg.get("To", ""))
    cc_addr = extract_addresses(msg.get("Cc", ""))
    subject = decode_mime_header(msg.get("Subject", ""))

    date_raw = msg.get("Date", "")
    try:
        dt = parsedate_to_datetime(date_raw)
        dt_cst = dt.astimezone(CST)
        date_str = dt_cst.strftime("%Y/%m/%d %H:%M:%S")
    except Exception:
        date_str = date_raw.strip()

    body = extract_body(msg)[:8000]

    return {
        "file": filename,
        "from": from_addr,
        "date": date_str,
        "to": to_addr,
        "cc": cc_addr,
        "subject": subject,
        "body": body,
    }


# ===================== Excel 导出 =====================

def build_excel_bytes(records: list) -> bytes:
    """构建 Excel 并返回 bytes"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"

    COL_WIDTHS = [55, 38, 22, 50, 38, 60, 100]

    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    data_font = Font(name="Microsoft YaHei", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    date_align = Alignment(horizontal="center", vertical="top")

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for col_idx, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, record in enumerate(records, 2):
        for col_idx, key in enumerate(HEADERS, 1):
            value = record.get(key, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = date_align if key == "date" else data_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(records) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ===================== UI =====================

st.title("📧 Nexas EML 邮件解析器")
st.caption("上传 .eml 文件 → 自动解析 → 交互浏览 → 导出 Excel")

# --- 侧边栏 ---
with st.sidebar:
    st.header("⚙️ 操作区")

    uploaded_files = st.file_uploader(
        "拖拽或点击上传 .eml 文件",
        type=["eml"],
        accept_multiple_files=True,
        help="支持批量上传，单次建议不超过 500 个文件",
    )

    # 示例数据生成按钮
    st.divider()
    if st.button("🧪 生成示例 EML 并测试", use_container_width=True):
        # 构造一个简单 EML
        sample_eml = (
            "From: sender@example.com\r\n"
            "To: receiver@example.com\r\n"
            "Cc: cc@example.com\r\n"
            "Date: Tue, 15 Jan 2025 14:30:00 +0900\r\n"
            "Subject: =?utf-8?B?5rWL6K+V6YKu5Lu2?=\r\n"
            "Content-Type: text/plain; charset=utf-8\r\n"
            "\r\n"
            "这是一封测试邮件。\r\n"
            "用于验证 EML 解析器功能。\r\n"
        )
        st.session_state["sample_data"] = sample_eml
        st.success("示例数据已生成，点击下方按钮加载")

    if "sample_data" in st.session_state and st.button("📥 加载示例数据", use_container_width=True):
        if "uploaded_files_processed" not in st.session_state:
            st.session_state["uploaded_files_processed"] = False
        # 将示例数据作为伪上传文件
        st.session_state["_sample_trigger"] = True
        st.rerun()

    st.divider()
    st.markdown("### 📊 项目信息")
    st.markdown("- 仓库: [AllenZorest/EML-to-EXCEL](https://github.com/AllenZorest/EML-to-EXCEL)")
    st.markdown("- 技术栈: Python + Streamlit + openpyxl")
    st.markdown("- 支持中/日/英多语言邮件")

# --- 主区域 ---
if uploaded_files or st.session_state.get("_sample_trigger"):
    # 处理上传文件
    if "results" not in st.session_state or not st.session_state.get("_processed"):
        results = []
        success = 0
        failed = 0

        progress_bar = st.progress(0, text="正在解析...")
        status_text = st.empty()

        files_to_process = uploaded_files if uploaded_files else []

        # 示例数据模式
        if st.session_state.get("_sample_trigger") and not uploaded_files:
            sample_eml = st.session_state.get("sample_data", "")
            record = parse_eml_bytes("test_sample.eml", sample_eml.encode("utf-8"))
            if record:
                results.append(record)
                success += 1
            else:
                failed += 1
            st.session_state["_sample_trigger"] = False

        total = len(files_to_process)
        for i, f in enumerate(files_to_process):
            data = f.getvalue()
            record = parse_eml_bytes(f.name, data)
            if record:
                results.append(record)
                success += 1
            else:
                failed += 1

            pct = int((i + 1) / total * 100) if total else 100
            progress_bar.progress(pct, text=f"解析中... {i+1}/{total}")
            status_text.text(f"成功: {success}  |  失败: {failed}")

        progress_bar.empty()
        status_text.empty()

        st.session_state["results"] = results
        st.session_state["success_count"] = success
        st.session_state["failed_count"] = failed
        st.session_state["_processed"] = True

    results = st.session_state.get("results", [])
    success = st.session_state.get("success_count", 0)
    failed = st.session_state.get("failed_count", 0)

    # --- 统计卡片 ---
    st.markdown("---")
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("📬 总文件数", len(results) + failed)
    col2.metric("✅ 成功解析", success, delta=None)
    col3.metric("❌ 解析失败", failed, delta=None, delta_color="inverse")
    col4.metric("📊 成功率", f"{success/(success+failed)*100:.0f}%" if (success + failed) > 0 else "N/A")

    # 日期范围
    df = pd.DataFrame(results, columns=HEADERS) if results else pd.DataFrame()
    if not df.empty and "date" in df.columns:
        try:
            dates = pd.to_datetime(df["date"], errors="coerce").dropna()
            if not dates.empty:
                date_range = f"{dates.min().strftime('%m/%d')} – {dates.max().strftime('%m/%d')}"
            else:
                date_range = "N/A"
        except Exception:
            date_range = "N/A"
    else:
        date_range = "N/A"
    col5.metric("📅 日期范围", date_range)

    # --- 操作按钮 ---
    col_a, col_b, col_c = st.columns([1, 1, 3])
    with col_a:
        if st.button("🔄 重新解析", use_container_width=True):
            for key in ["results", "success_count", "failed_count", "_processed"]:
                st.session_state.pop(key, None)
            st.rerun()

    with col_b:
        if results:
            excel_data = build_excel_bytes(results)
            st.download_button(
                label="📥 导出 Excel",
                data=excel_data,
                file_name=f"emails_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # --- 数据表格 ---
    if results:
        st.markdown("---")
        st.subheader(f"📋 解析结果（共 {len(results)} 条）")

        # 搜索/筛选栏
        search_col1, search_col2 = st.columns([3, 1])
        with search_col1:
            search_term = st.text_input("🔍 搜索（标题/发件人/正文）", placeholder="输入关键词筛选...")

        # 过滤
        display_df = df.copy()
        if search_term:
            mask = (
                display_df["subject"].str.contains(search_term, case=False, na=False)
                | display_df["from"].str.contains(search_term, case=False, na=False)
                | display_df["body"].str.contains(search_term, case=False, na=False)
            )
            display_df = display_df[mask]

        with search_col2:
            st.caption(f"显示 {len(display_df)} / {len(df)} 条")

        # 表格渲染
        st.dataframe(
            display_df[["file", "from", "date", "to", "subject"]],
            use_container_width=True,
            height=500,
            column_config={
                "file": st.column_config.TextColumn("文件名", width="medium"),
                "from": st.column_config.TextColumn("发件人", width="medium"),
                "date": st.column_config.TextColumn("日期", width="small"),
                "to": st.column_config.TextColumn("收件人", width="medium"),
                "subject": st.column_config.TextColumn("标题", width="large"),
            },
            hide_index=True,
        )

        # --- 发件人统计 ---
        st.markdown("---")
        st.subheader("📊 发件人分布")

        sender_counts = Counter(r.get("from", "未知") for r in results)
        top_senders = sender_counts.most_common(10)

        if top_senders:
            sender_df = pd.DataFrame(top_senders, columns=["发件人", "邮件数"])
            col_chart, col_table = st.columns([1, 1])
            with col_chart:
                st.bar_chart(sender_df.set_index("发件人"), use_container_width=True)
            with col_table:
                st.dataframe(sender_df, hide_index=True, use_container_width=True)

        # --- 正文预览 ---
        st.markdown("---")
        st.subheader("📝 邮件正文预览")

        selected_file = st.selectbox(
            "选择邮件查看正文",
            options=[r["file"] for r in results],
            format_func=lambda x: f"{x}",
        )

        if selected_file:
            selected_record = next((r for r in results if r["file"] == selected_file), None)
            if selected_record:
                col_info, col_body = st.columns([1, 2])
                with col_info:
                    st.markdown("**邮件信息**")
                    st.markdown(f"**发件人:** {selected_record['from']}")
                    st.markdown(f"**收件人:** {selected_record['to']}")
                    if selected_record["cc"]:
                        st.markdown(f"**抄送:** {selected_record['cc']}")
                    st.markdown(f"**日期:** {selected_record['date']}")
                    st.markdown(f"**标题:** {selected_record['subject']}")

                with col_body:
                    st.markdown("**正文内容**")
                    body_text = selected_record["body"]
                    if len(body_text) > 3000:
                        preview = body_text[:3000] + f"\n\n... (共 {len(body_text)} 字符，已截断)"
                    else:
                        preview = body_text
                    st.text_area(
                        label="",
                        value=preview,
                        height=300,
                        disabled=True,
                        label_visibility="collapsed",
                    )

    else:
        st.info("解析结果为空，请检查上传的 EML 文件格式。")
else:
    # 初始状态 — 上传引导
    st.markdown("### 👋 使用方式")
    st.markdown("""
    1. **上传文件** — 在左侧边栏拖拽或选择 .eml 文件
    2. **自动解析** — 提取发件人、收件人、标题、日期、正文
    3. **交互浏览** — 表格筛选、搜索、正文预览
    4. **导出结果** — 一键下载为格式化 Excel
    
    > 💡 点击侧边栏「生成示例 EML 并测试」可快速体验功能
    """)

    st.info("📂 请在左侧边栏上传 .eml 文件开始解析")
